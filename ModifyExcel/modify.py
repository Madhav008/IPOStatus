import openpyxl
import pandas as pd
import concurrent.futures
import os

def read_excel_in_chunks(input_excel_file, chunk_size):
    wb = openpyxl.load_workbook(input_excel_file, read_only=True)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        # Get column names from the first row
        columns = [cell.value for cell in sheet[1]]

        for row in range(2, sheet.max_row + 1, chunk_size):
            rows = sheet.iter_rows(min_row=row, max_row=min(row + chunk_size - 1, sheet.max_row), values_only=True)
            chunk_data = [row_data for row_data in rows]
            yield sheet_name, pd.DataFrame(chunk_data, columns=columns)

def process_chunk(sheet_name, chunk, output_folder, company_name_value, counter):
    chunk['Company_Name'] = company_name_value
    csv_file_path = f"{output_folder}/{sheet_name}_{counter}.csv"
    chunk.to_csv(csv_file_path, index=False)
    print(f"Sheet '{sheet_name}' saved as CSV: {csv_file_path}")

def split_excel_to_csv(input_excel_file, output_folder, company_name_value, chunk_size=10000):
    # Use ThreadPoolExecutor for parallel processing
    with concurrent.futures.ThreadPoolExecutor() as executor:
        # Submit each sheet processing task to the executor
        futures = []
        counter = 1

        for sheet_name, chunk in read_excel_in_chunks(input_excel_file, chunk_size):
            # Submit each task with a unique counter
            futures.append(executor.submit(process_chunk, sheet_name, chunk, output_folder, company_name_value, counter))
            counter += 1

        # Wait for all tasks to complete
        concurrent.futures.wait(futures)

# Specify the path to the Excel file, the output folder for CSV files, and the Company_Name value
input_excel_file = "/mnt/c/Users/madha/Downloads/DOMS_ALL.xlsx"
output_folder = "Files"
company_name_value = "DOMS"

# Create the output folder if it doesn't exist
os.makedirs(output_folder, exist_ok=True)

# Call the function to split the Excel file into CSV files with chunking and threading
split_excel_to_csv(input_excel_file, output_folder, company_name_value)
